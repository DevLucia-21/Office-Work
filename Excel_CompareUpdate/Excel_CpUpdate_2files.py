from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

# 엑셀 파일 경로 설정
file1 = r"파일경로\파일명.xlsx"
file2 = r"파일경로\파일명.xlsx"

# 엑셀 파일 로드
wb1 = load_workbook(file1)
wb2 = load_workbook(file2)

ws1 = wb1["엑셀1의 시트명"]
ws2 = wb2["엑셀2의 시트명"]

# 색상 설정
green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # 초록색 (값이 같을 때)
orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # 주황색 (값이 다를 때)

# B14:B255 범위 확인 (엑셀1)
for row1 in ws1.iter_rows(min_row=14, max_row=255, min_col=2, max_col=2):  # B열 (2번째 열)
    cell1 = row1[0]  # B열의 셀 값
    value1 = cell1.value

    if value1 is None:
        continue  # 빈 값이면 스킵

    # 엑셀2에서 같은 값을 A열에서 찾기
    found = False
    for row2 in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=1):  # A열 (엑셀2)
        cell2 = row2[0]
        if cell2.value == value1:  # 같은 값이 있으면 비교 시작
            found = True
            row_idx1 = cell1.row  # 엑셀1의 행 (H~S 비교용)
            row_idx2 = cell2.row  # 엑셀2의 행 (B~M 비교용)

            # H~S (엑셀1) ↔ B~M (엑셀2) 비교
            for col_offset in range(12):  # 12개의 컬럼 비교 (H~S ↔ B~M)
                cell1_target = ws1.cell(row=row_idx1, column=8 + col_offset)  # 엑셀1 H~S
                cell2_target = ws2.cell(row=row_idx2, column=2 + col_offset)  # 엑셀2 B~M

                if cell1_target.value == cell2_target.value:
                    cell1_target.fill = green_fill  # 값이 같으면 초록색
                else:
                    cell1_target.value = cell2_target.value  # 값이 다르면 엑셀2 값 복사
                    cell1_target.fill = orange_fill  # 주황색 변경
            break

    if not found:
        print(f"엑셀2에서 {value1}을(를) 찾을 수 없음 (B{cell1.row})")

# 수정된 파일 저장
file1_name = os.path.basename(file1)
save_path = os.path.join(os.path.dirname(file1), "updated_" + file1_name)

wb1.save(save_path)
wb1.close()
wb2.close()

print(f"엑셀 비교 및 색상 변경 완료! 파일 저장 위치: {save_path}")
