import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill

# 엑셀 파일 및 시트 지정(\역슬래시 사용 시 \u가 유니코드 이스케이프 시퀀스로 인식되어 에러나서 /슬래시 사용)
file_path = "파일경로/파일명.xlsx"  # 파일 경로 수정
sheet_name = "시트명"  # 원하는 시트명으로 변경

# 엑셀 파일 불러오기
wb = openpyxl.load_workbook(file_path)
ws = wb[sheet_name]  # 특정 시트 선택

# 색상 정의 (사용자가 지정한 색상)
green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # 연두색
orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # 주황색

# 값 변환 함수
def clean_value(value):
    """ 값의 형식을 정리하여 비교할 수 있도록 변환 """
    if value is None:
        return None  # None 값 유지
    
    # 문자열로 변환 후 공백/특수문자 제거
    str_value = str(value).strip().replace("\xa0", "").replace("\u200b", "")

    # 숫자로 변환 시도 (엑셀 형식 문제 해결)
    try:
        num_value = pd.to_numeric(str_value, errors='coerce')  # 숫자로 변환, 실패 시 NaN 반환
        if pd.notna(num_value):  # NaN이 아니면 숫자로 변환 성공
            return round(num_value, 6)  # 소수점 6자리까지 반올림 (오차 방지)
    except:
        pass
    
    return str_value.lower()  # 변환 실패 시 문자열로 비교

# 비교 및 처리 함수
def compare_and_update(start_col_1, start_col_2):
    for row in range(8, 374):  # 범위: 8~373
        for col_offset in range(3):  # 3개의 컬럼 그룹 처리 (0,1,2)
            col1 = start_col_1 + col_offset
            col2 = start_col_2 + col_offset

            cell1 = ws.cell(row=row, column=col1)
            cell2 = ws.cell(row=row, column=col2)

            value1 = clean_value(cell1.value)
            value2 = clean_value(cell2.value)

            if value1 == value2:
                cell1.fill = green_fill  # 연두색으로 채색
            else:
                cell1.value = cell2.value  # 값 복사
                cell1.fill = orange_fill  # 주황색으로 채색

# C~E vs F~H 비교
compare_and_update(3, 6)

# M~O vs P~R 비교
compare_and_update(13, 16)

# 변경 사항 저장
wb.save("updated_excel2.xlsx")  # 새로운 파일로 저장
wb.close()
