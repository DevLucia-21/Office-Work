import os
import openpyxl
from openpyxl.styles import PatternFill

# 파일 경로 직접 할당
excel1_path = r"파일경로\파일명.xlsx"
excel1_sheet_name = "시트명"

excel2_path = r"파일경로\파일명.xlsx"

# 엑셀 파일 열기
wb1 = openpyxl.load_workbook(excel1_path)
ws1 = wb1[excel1_sheet_name]

wb2 = openpyxl.load_workbook(excel2_path)
ws2 = wb2.active  # 엑셀2는 첫 번째(유일한) 시트를 자동 선택

# 색상 채우기 정의
green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

# 엑셀1의 15행부터 실행
for row in ws1.iter_rows(min_row=15):
    row_number = row[0].row  # 현재 행 번호

    x_value = ws1[f"U{row_number}"].value  # 엑셀1의 X열 값
    if x_value is None:
        continue

    # 엑셀2의 C열에서 x_value와 일치하는 행 찾기
    matched_row_number = None
    for cell in ws2["C"]:
        if cell.value == x_value:
            matched_row_number = cell.row
            break

    if matched_row_number is None:
        continue  # 매칭되는 값이 없으면 다음 행으로

    bx_value = ws2[f"BY{matched_row_number}"].value  # 엑셀2의 BX열 값
    k_cell = ws1[f"I{row_number}"]  # 엑셀1의 K열 값
    k_value = k_cell.value

    # 비교 후 셀 값 및 색상 변경
    if k_value == bx_value:
        k_cell.fill = green_fill  # 값이 같으면 연두색
    else:
        k_cell.value = bx_value  # 값이 다르면 엑셀2 값 복사
        k_cell.fill = orange_fill  # 주황색

# 원래 파일명에서 폴더와 확장자 분리
folder, original_filename = os.path.split(excel1_path)
filename_without_ext, ext = os.path.splitext(original_filename)

# "Updated_" 붙여서 새 파일명 생성
new_filename = os.path.join(folder, f"Updated_{filename_without_ext}{ext}")

# 수정된 엑셀 저장
wb1.save(new_filename)
print(f"작업이 완료되었습니다. 파일이 저장됨: {new_filename}")
