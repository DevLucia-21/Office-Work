import openpyxl
from openpyxl.styles import PatternFill

# 엑셀 파일 및 시트 지정(\역슬래시 사용 시 \u가 유니코드 이스케이프 시퀀스로 인식되어 에러나서 /슬래시 사용)
file_path = "파일경로/파일명.xlsx"  # 파일 경로 수정
sheet_name = "시트명"  # 원하는 시트명으로 변경

# 엑셀 파일 불러오기
wb = openpyxl.load_workbook(file_path)
ws = wb[sheet_name]  # 특정 시트 선택

# 색상 정의 (사용자가 지정한 색상)
green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # 연두색
orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # 주황색

# 비교 및 처리 함수
def compare_and_update(start_col_1, start_col_2):
    for row in range(8, 374):  # 범위: 8~373
        for col_offset in range(3):  # 3개의 컬럼 그룹 처리 (0,1,2)
            col1 = start_col_1 + col_offset
            col2 = start_col_2 + col_offset

            cell1 = ws.cell(row=row, column=col1)
            cell2 = ws.cell(row=row, column=col2)

            # None 값을 빈 문자열로 변환
            value1 = cell1.value if cell1.value is not None else ""
            value2 = cell2.value if cell2.value is not None else ""

            # 숫자로 변환 시도 (문자면 그대로 비교)
            def to_number(value):
                try:
                    return float(value)  # 정수/소수 변환
                except ValueError:
                    return str(value).strip().lower()  # 변환 불가한 경우 문자열로 처리

            value1 = to_number(value1)
            value2 = to_number(value2)

            if value1 == value2:
                cell1.fill = green_fill  # 연두색으로 채색
            else:
                cell1.value = cell2.value  # 값 복사
                cell1.fill = orange_fill  # 주황색으로 채색

# C~E vs F~H 비교
compare_and_update(3, 6)

# M~O vs P~R 비교
compare_and_update(13, 16)

# 변경 사항 저장
wb.save("updated_excel.xlsx")  # 새로운 파일로 저장
wb.close()
